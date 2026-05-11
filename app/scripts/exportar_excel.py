import os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR   = os.path.dirname(SCRIPT_DIR)
DATA_DIR   = os.path.join(BASE_DIR, "data")
CSV_IN     = os.path.join(DATA_DIR, "timeout_history.csv")
XLSX_OUT   = os.path.join(DATA_DIR, "Historico_Timeouts.xlsx")

def leer_csv():
    if not os.path.exists(CSV_IN):
        sys.exit(f"No se encontro el CSV: {CSV_IN}")
    for sep in [",", ";"]:
        for enc in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                df = pd.read_csv(CSV_IN, sep=sep, encoding=enc)
                if len(df.columns) > 1:
                    return df
            except Exception:
                continue
    sys.exit("No se pudo leer el CSV.")

df = leer_csv()

def reparar(s):
    if not isinstance(s, str): return s
    if "Ã" in s:
        try: return s.encode("latin-1").decode("utf-8")
        except: return s
    return s

for col in df.select_dtypes(include="object").columns:
    df[col] = df[col].apply(reparar)

df["fecha_deteccion"] = pd.to_datetime(df["fecha_deteccion"], errors="coerce")
df["fecha_timeout"]   = pd.to_datetime(df["fecha_timeout"],   errors="coerce")
df = df.sort_values("fecha_timeout", ascending=False).reset_index(drop=True)
df["time_taken_s"]    = (pd.to_numeric(df["time_taken_ms"], errors="coerce") / 1000).round(2)

print(f"Registros: {len(df):,} | CLARO: {(df['operador']=='CLARO').sum()} | ETB: {(df['operador']=='ETB').sum()}")

HF = PatternFill("solid", start_color="1A1A2E")
HFN= Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CF = PatternFill("solid", start_color="FFE5E5")
EF = PatternFill("solid", start_color="DCEBFD")
TF = PatternFill("solid", start_color="FFF3C7")
TTL= Font(name="Calibri", size=18, bold=True, color="1A1A2E")
SUB= Font(name="Calibri", size=12, bold=True, color="333333")
NF = Font(name="Calibri", size=11)
BF = Font(name="Calibri", size=11, bold=True)
CT = Alignment(horizontal="center", vertical="center")
LF = Alignment(horizontal="left", vertical="center")
BD = Border(left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"))

wb = Workbook()
ws = wb.active; ws.title = "Resumen"
ws["A1"] = "HISTORICO DE TIMEOUTS - PORTAL NOC ATP"
ws["A1"].font = TTL; ws.merge_cells("A1:E1"); ws.row_dimensions[1].height = 30
ws["A2"] = f"Generado: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')}"
ws["A2"].font = Font(size=10, italic=True, color="888888"); ws.merge_cells("A2:E2")
ws["A4"] = "INDICADORES PRINCIPALES"; ws["A4"].font = SUB

kpis = [
    ("Total timeouts:", len(df)),
    ("CLARO:", int((df["operador"]=="CLARO").sum())),
    ("ETB:", int((df["operador"]=="ETB").sum())),
    ("Tiempo promedio (s):", round(df["time_taken_s"].mean(),2)),
    ("Tiempo maximo (s):", round(df["time_taken_s"].max(),2)),
    ("Procesos distintos:", int(df["proceso"].nunique())),
    ("Primer evento:", df["fecha_timeout"].min().strftime("%d/%m/%Y %H:%M") if pd.notna(df["fecha_timeout"].min()) else "-"),
    ("Ultimo evento:", df["fecha_timeout"].max().strftime("%d/%m/%Y %H:%M") if pd.notna(df["fecha_timeout"].max()) else "-"),
]
r=5
for e,v in kpis:
    ws.cell(row=r, column=1, value=e).font = BF
    c = ws.cell(row=r, column=2, value=v); c.font = NF
    if isinstance(v,(int,float)): c.number_format = "#,##0.00" if isinstance(v,float) else "#,##0"
    r+=1

r+=2
ws.cell(row=r, column=1, value="TOP 10 PROCESOS").font = SUB
r+=1
for j,h in enumerate(["Proceso","CLARO","ETB","Total"],1):
    c=ws.cell(row=r,column=j,value=h); c.font=HFN; c.fill=HF; c.alignment=CT; c.border=BD
top = df.groupby(["proceso","operador"]).size().unstack(fill_value=0)
if "CLARO" not in top.columns: top["CLARO"]=0
if "ETB" not in top.columns: top["ETB"]=0
top["Total"] = top["CLARO"]+top["ETB"]
top = top.sort_values("Total", ascending=False).head(10)
r+=1
for p,f in top.iterrows():
    ws.cell(row=r,column=1,value=str(p)).font=NF
    ws.cell(row=r,column=2,value=int(f["CLARO"])).font=NF
    ws.cell(row=r,column=3,value=int(f["ETB"])).font=NF
    ws.cell(row=r,column=4,value=int(f["Total"])).font=BF
    for ci in range(1,5):
        c=ws.cell(row=r,column=ci); c.alignment=LF if ci==1 else CT; c.border=BD
        if ci==4: c.fill=TF
    r+=1

ws.column_dimensions["A"].width=42
for L in ["B","C","D","E"]: ws.column_dimensions[L].width=18

ws2 = wb.create_sheet("Detalle Timeouts")
cols  = ["fecha_timeout","operador","proceso","codigo","descripcion","time_taken_ms","time_taken_s","status","status_code","rango_consulta","fecha_deteccion"]
heads = ["Fecha Timeout","Operador","Proceso","Codigo","Descripcion","Tiempo (ms)","Tiempo (s)","Estado","HTTP","Origen","Fecha Deteccion"]
for j,h in enumerate(heads,1):
    c=ws2.cell(row=1,column=j,value=h); c.font=HFN; c.fill=HF; c.alignment=CT; c.border=BD
for i,reg in enumerate(df[cols].itertuples(index=False), start=2):
    vals=list(reg); op=vals[1]
    fill = CF if op=="CLARO" else (EF if op=="ETB" else None)
    for j,v in enumerate(vals,1):
        c=ws2.cell(row=i,column=j); c.value="" if pd.isna(v) else v; c.font=NF; c.border=BD
        if fill: c.fill=fill
        if j in (1,11) and not pd.isna(v): c.number_format="dd/mm/yyyy hh:mm:ss"; c.alignment=CT
        elif j in (2,6,7,8,9): c.alignment=CT
        else: c.alignment=LF
        if j==6 and isinstance(v,(int,float)) and not pd.isna(v): c.number_format="#,##0"

ult = len(df)+1
tab = Table(displayName="TablaTimeouts", ref=f"A1:{get_column_letter(len(heads))}{ult}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
ws2.add_table(tab)

anchos = {"A":20,"B":11,"C":26,"D":22,"E":40,"F":14,"G":12,"H":12,"I":9,"J":22,"K":20}
for col,w in anchos.items(): ws2.column_dimensions[col].width=w
ws2.freeze_panes="A2"; ws2.row_dimensions[1].height=22

ws2.conditional_formatting.add(f"F2:F{ult}", CellIsRule(operator="greaterThan", formula=["40000"],
    fill=PatternFill("solid", start_color="FCA5A5"),
    font=Font(name="Calibri", size=11, bold=True, color="7F1D1D")))

os.makedirs(DATA_DIR, exist_ok=True)
wb.save(XLSX_OUT)
print(f"\nOK -> {XLSX_OUT}")
